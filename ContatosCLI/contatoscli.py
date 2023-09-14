import openpyxl
import os

def cadastrar_registro(sheet):
    nome = input("Digite o nome: ")
    empresa = input("Digite a empresa: ")
    telefone = input("Digite o telefone (com DDD): ")
    
    # Armazena os dados em uma lista
    registro = [nome, empresa, telefone]
    
    # Adiciona o registro à planilha
    sheet.append(registro)
    
    # Imprime a pré-planilha
    imprimir_planilha(sheet)

    print("Registro adicionado com sucesso!")

def editar_registro(sheet):
    numero_linha = int(input("Digite o número da linha que deseja editar: "))
    
    if 1 <= numero_linha <= sheet.max_row:
        nome = input("Digite o novo nome: ")
        empresa = input("Digite a nova empresa: ")
        telefone = input("Digite o novo telefone (com DDD): ")
        
        # Atualiza os dados na planilha
        sheet.cell(row=numero_linha, column=1, value=nome)
        sheet.cell(row=numero_linha, column=2, value=empresa)
        sheet.cell(row=numero_linha, column=3, value=telefone)

        # Imprime a pré-planilha
        imprimir_planilha(sheet)

        print("Registro atualizado com sucesso!")
    else:
        print("Linha não encontrada.")

def deletar_registro(sheet):
    numero_linha = int(input("Digite o número da linha que deseja deletar: "))
    
    if 1 <= numero_linha <= sheet.max_row:
        sheet.delete_rows(numero_linha)

        # Imprime a pré-planilha
        imprimir_planilha(sheet)

        print("Registro deletado com sucesso!")
    else:
        print("Linha não encontrada.")

def imprimir_planilha(sheet):
    print("\nPré-Planilha:")
    for row in sheet.iter_rows(values_only=True):
        print(" | ".join(map(str, row)))

# Verifica se o arquivo 'Contatos.xlsx' já existe
if os.path.exists("Contatos.xlsx"):
    # Se existe, abre o arquivo existente
    workbook = openpyxl.load_workbook("Contatos.xlsx")
    sheet = workbook.active
else:
    # Se não existe, cria um novo arquivo e adiciona um cabeçalho
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Contatos"
    sheet.append(["Nome", "Empresa", "Telefone (com DDD)"])

# Seção "Sobre" a imprimir no início do programa
print("ContatosCLI v2.0 - Copyright (C) Bernardo Krzysczak <bernardokrz@gmail.com>")
print("EN: This software is licensed under the GNU General Public License v3.0 and provided as is. Any forks and/or derivatives must reference this project with the author's name and contact.")

print("BR: Este software é licenciado sob a GNU General Public License v3.0 e disponibilizado no estado em que se encontra. Quaisquer forks e/ou derivados devem referenciar este projeto com o nome do autor e contato.")
while True:
    print("\nMenu:")
    print("1 - Cadastrar registro")
    print("2 - Editar registro")
    print("3 - Deletar registro")
    print("4 - Sair")
    
    opcao = input("Escolha uma opção: ")
    
    if opcao == '1':
        cadastrar_registro(sheet)
    elif opcao == '2':
        editar_registro(sheet)
    elif opcao == '3':
        deletar_registro(sheet)
    elif opcao == '4':
        # Salva a planilha e sai do programa
        workbook.save("Contatos.xlsx")
        print("Planilha Excel gerada/atualizada com sucesso.")
        print("Saindo do programa.")
        break
    else:
        print("Opção inválida. Por favor, escolha uma opção válida.")
