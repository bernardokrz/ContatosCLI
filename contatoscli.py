# Exportar dados cadastrados para Excel
from openpyxl import Workbook
def exportar_para_excel():
    # Ordena os contatos em ordem alfabética pelo nome
    contatos_ordenados = sorted(contact.keys())

    # Cria um novo workbook
    workbook = Workbook()
            
    # Seleciona a planilha ativa
    sheet = workbook.active
            
    # Escreve o cabeçalho
    sheet['A1'] = 'Nome'
    sheet['B1'] = 'Telefone'
            
    # Escreve os contatos em ordem alfabética
    for i, nome in enumerate(contatos_ordenados, start=2):
        telefone = contact[nome]
        sheet.cell(row=i, column=1, value=nome)
        sheet.cell(row=i, column=2, value=telefone)

    # Gera o timestamp para o nome do arquivo
    import datetime
    timestamp = datetime.datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
    nome_arquivo = f"contatos_{timestamp}.xlsx"
    
    # Salva o arquivo Excel
    workbook.save(nome_arquivo)

# Cria o dicionário de contatos
contact = {}
# Função que mostra lista de contatos cadastrados em ordem alfabética
def mostrar_contato():
    print("Nome\t\t\tTelefone")
    for key in sorted(contact.keys()):
        print("{}\t\t{}".format(key, contact.get(key)))


# Mensagem de inicialização
print("<ContatosCLI> Versão 1.2\npor Bernardo Krzysczak - krz02@proton.me\nSOFTWARE LIVRE PARA USO PESSOAL E COMERCIAL\n")

# Menu principal
while True:
    escolha = int(input("1. Novo contato\n2. Pesquisar contato\n3. Exibir contatos cadastrados\n4. Editar contato\n5. Apagar contato\n6. Sair\n\nEscolha a opção desejada: "))

# Opções do menu
    if escolha == 1:

        nome = input("Digite o nome do contato: ")
        telefone = input("Digite o telefone com DDD: ")
        contact[nome] = telefone
        print("\n**Contato cadastrado.\n")
    elif escolha == 2:
        procurar_nome = input("Digite o nome do contato: ")
        if procurar_nome in contact:
            print("\n**O número de",procurar_nome,"é:", contact[procurar_nome], "\n")
        else:
            print("\n**Contato não encontrado.")
    elif escolha == 3:
        if not contact:
            print("\n**Lista de contatos vazia!")
        else:
            print("-------------CONTATOS CADASTRADOS (nesta sessão)-------------")
            mostrar_contato()
    elif escolha == 4:
        editar_contato = input("Digite o nome do contato a ser editado: ")
        if editar_contato in contact:
            telefone = input("Digite o novo telefone com DDD: ")
            contact[editar_contato]=telefone
            print("\n**Contato atualizado.\n")
            mostrar_contato()
        else:
            print("\n**Contato não encontrado.")
    elif escolha == 5:
        apagar_contato = input("Digite o nome do contato a ser apagado: ")
        if apagar_contato in contact:
            confirmar = input("Tem certeza que deseja apagar este contato?(s/n): ")
            if confirmar == 's' or confirmar =='S':
                contact.pop(apagar_contato)
                print("\n**Contato apagado.\n")
            mostrar_contato()
        else:
            print("\n**Contato não encontrado.")
# Encerrando o programa e gerando arquivo Excel:
    else:
        print("**Encerrando...")
        exportar_para_excel()
        break